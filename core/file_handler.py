import pandas as pd
import io
import logging
import os
from pathlib import Path
from openpyxl.styles import PatternFill

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None

logger = logging.getLogger(__name__)

def read_command_file(filepath: str) -> list[str]:
    """명령어 파일(txt/xlsx)을 읽어 순서대로 반환합니다."""
    path = Path(filepath)
    suffix = path.suffix.lower()
    commands: list[str] = []

    try:
        if suffix in (".xlsx", ".xls", ".xlsm"):
            df = pd.read_excel(filepath, header=None)
            if df.empty:
                return []
            first_col = df.iloc[:, 0].tolist()
            for value in first_col:
                if pd.isna(value):
                    continue
                line = str(value).strip()
                if line:
                    commands.append(line)
        elif suffix == ".txt":
            with open(filepath, "r", encoding="utf-8") as f:
                for line in f:
                    cleaned = line.strip()
                    if cleaned:
                        commands.append(cleaned)
        else:
            raise ValueError(f"지원하지 않는 파일 형식입니다: {suffix}")
    except Exception as e:
        logger.error("명령어 파일 읽기 실패: %s, 오류: %s", filepath, e)
        raise

    return commands

def read_excel_file(filepath: str, password: str = None) -> pd.DataFrame:
    """
    엑셀 파일을 읽어 데이터프레임으로 반환합니다.
    암호화된 경우 암호를 사용하여 복호화합니다.
    """
    try:
        if password:
            if msoffcrypto is None:
                raise ImportError("암호화된 Excel 파일 지원을 위해 'msoffcrypto-tool' 라이브러리를 설치해주세요.")
            
            decrypted_file = io.BytesIO()
            with open(filepath, 'rb') as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                office_file.decrypt(decrypted_file)
            
            df = pd.read_excel(decrypted_file)
            logger.info("암호화된 파일 복호화 성공: %s", filepath)
        else:
            df = pd.read_excel(filepath)
            logger.info("선택한 파일: %s", filepath)
        
        return df
    except Exception as e:
        logger.error("엑셀 파일 읽기 실패: %s, 오류: %s", filepath, e)
        raise

def save_results_to_excel(
    results: list,
    output_filepath: str,
    column_order: list[str] | None = None
):
    """결과를 엑셀 파일에 저장하고, 실패한 항목에 서식을 적용합니다."""
    try:
        logger.info("결과 저장 시작: %s", output_filepath)

        output_dir = os.path.dirname(output_filepath)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        
        processed_results = []
        for res in results:
            row = {
                'ip': res.get('ip'),
                'vendor': res.get('vendor'),
                'os': res.get('os'),
                '접속 상태': '성공' if res.get('status') == 'success' else '실패',
                '오류 메시지': res.get('error_message', '')
            }
            if res.get('inspection_results'):
                for key, value in res['inspection_results'].items():
                    if not key.startswith('error_') and key not in ['error', 'backup_error', 'backup_file']:
                        row[key] = value
            
            processed_results.append(row)

        if not processed_results:
            logger.warning("저장할 결과가 없습니다.")
            return

        df = pd.DataFrame(processed_results)
        
        base_cols = ['ip', 'vendor', 'os', '접속 상태', '오류 메시지']
        if column_order:
            ordered_inspection_cols = [
                col for col in column_order
                if col in df.columns and col not in base_cols
            ]
            remaining_cols = [
                col for col in df.columns
                if col not in base_cols and col not in ordered_inspection_cols
            ]
            df = df[base_cols + ordered_inspection_cols + remaining_cols]
        else:
            other_cols = [col for col in df.columns if col not in base_cols]
            df = df[base_cols + other_cols]

        with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Inspection Results')
            
            worksheet = writer.sheets['Inspection Results']
            
            light_red_fill = PatternFill(start_color='FFFFC7CE',
                                         end_color='FFFFC7CE',
                                         fill_type='solid')
            
            for idx, row in df.iterrows():
                if row['접속 상태'] == '실패':
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=idx + 2, column=col_idx).fill = light_red_fill
                        
            for column_cells in worksheet.columns:
                try:
                    length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
                except (ValueError, TypeError):
                    pass

        logger.info("결과가 저장되었습니다: %s", output_filepath)
        
    except Exception as e:
        logger.error("결과 저장 중 오류 발생: %s", e)
        raise 